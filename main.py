import openpyxl
import openpyxl as xl

from openpyxl import Workbook

INPUT_PATH = "C:\\Users\\baong\\OneDrive\\Desktop\\excel python copy\\Simulation - INPUT file.xlsx"
OUTPUT_PATH = "C:\\Users\\baong\\OneDrive\\Desktop\\excel python copy\\Simulation - OUTPUT file.xlsx"
ARCHIVE_FOLDER_PATH = "C:\\Users\\baong\\OneDrive\\Desktop\\excel python copy\\Versions archive\\"
SHEET_NAME = 'Sheet1'


def load_input_worksheet():
    input_workbook = xl.load_workbook(INPUT_PATH)
    input_worksheet = input_workbook[SHEET_NAME]
    return input_worksheet


def archive(version):
    # creating new blank excel file
    new_workbook = Workbook()
    new_name = "Simulation - INPUT file version" + version + ".xlsx"
    destination = ARCHIVE_FOLDER_PATH + new_name

    # loading from input excel
    input_worksheet = load_input_worksheet()

    # copying to new blank excel file
    new_worksheet = new_workbook[SHEET_NAME]

    for row in input_worksheet:
        for cell in row:
            new_worksheet[cell.coordinate].value = cell.value

    # change the version cell
    new_worksheet['A1'].value = version
    new_workbook.save(destination)


def find_starting_cell(output_version_cell, output_worksheet, output_version_column, version):
    chosen_cell = ''
    for number in range(output_version_cell.row + 1, output_worksheet.max_row):
        chosen_cell = output_worksheet.cell(number, output_version_column)

        # case 1 where we found a cell with the same version
        if chosen_cell.value == int(version):
            return chosen_cell
    # case 2 where we did not find a cell with the same version, and we return the last cell in the column
    return chosen_cell

 #todo change as value

def copy_data(version):
    # loading from input excel
    input_worksheet = load_input_worksheet()
    input_version_cell = get_version_cell(input_worksheet)
    input_version_cell_below = input_worksheet.cell(input_version_cell.row +1, input_version_cell.column) # eliminate header row

    max_cell = input_worksheet.cell(input_worksheet.max_column, input_worksheet.max_row)
    cell_range = input_worksheet[input_version_cell_below.coordinate :max_cell.coordinate]

    # find version cell in output to get the version column
    output_workbook = xl.load_workbook(OUTPUT_PATH)
    output_worksheet = output_workbook[SHEET_NAME]
    output_version_cell = get_version_cell(output_worksheet)
    output_version_column = output_version_cell.column

    # find location to paste to
    chosen_cell = find_starting_cell(output_version_cell, output_worksheet, output_version_column, version)

    # Paste cell range
    row_counter = 0
    for row in cell_range:
        cell_counter = 0
        for cell in row:
            if cell_counter == 0: # overwrite all value of version cells to indicated version
                output_worksheet.cell(chosen_cell.row + row_counter, chosen_cell.column).value = version
            else: # paste the rest of cell range
                output_worksheet.cell(chosen_cell.row + row_counter, chosen_cell.column + cell_counter).value = cell.value
            cell_counter += 1
        row_counter += 1
    # TODO: paste only value

    #  output_worksheet.cell(chosen_cell.row + row_counter, chosen_cell.column) <- specify coordinate of each cell

    output_workbook.save(OUTPUT_PATH)
    # TODO: instead of print, copy this to the output file with 2 scenarios: if find same version number then overwrite, if not, then write below where first blank


def get_version_cell(sheet):
    for row in sheet:
        for cell in row:
            value = str(cell.value)
            if value.lower() == "version":
                return cell
    raise ValueError('There is no specified cell')


# create: make a copy of the output file before changing anything to prevent accidental overwrites

def main():
    type = input("Do you want to archive file or copy data. Type 1 for archive, 2 for copy: ")
    version = input("Specify version number: ")
    if type == "1":
        archive(version)
    elif type == "2":
        copy_data(version)
    else:
        main()


# main method
if __name__ == '__main__':
    main()
